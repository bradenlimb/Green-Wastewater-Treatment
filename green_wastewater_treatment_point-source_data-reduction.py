#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue May 23 15:34:49 2023

@author: bradenlimb
"""

#%% Import Modules
from IPython import get_ipython
get_ipython().magic('reset -sf')
import pandas as pd
import geopandas as gp
import numpy as np
import sys
import os
import math
from tqdm import tqdm
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import pickle
import itertools
import datetime
begin_time = datetime.datetime.now()

#%% Set Important Assumptions for Analysis 

## TODO - Set the treatment level
min_treat_level = 5

input_data = {}
input_data['Level 2'] = {
    'conc_N': 8, # mg/L
    'conc_P': 1, # mg/L
    }
input_data['Level 3'] = {
    'conc_N': 6, # mg/L
    'conc_P': 0.2, # mg/L
    }
input_data['Level 4'] = {
    'conc_N': 3, # mg/L
    'conc_P': 0.1, # mg/L
    }
input_data['Level 5'] = {
    'conc_N': 2, # mg/L
    'conc_P': 0.02, # mg/L
    }

# Pull nutrient concentrations
treat_conc_N = input_data[f'Level {min_treat_level}']['conc_N'] # mg/L
treat_conc_P = input_data[f'Level {min_treat_level}']['conc_P'] # mg/L



farmer_incentive = 31 * 2.47105 # $/acre converted to $/ha - https://theoutcomesfund.com/news-release-2021-environmental-outcomes#:~:text=ANKENEY%2C%20IA%20and%20WASHINGTON%2C%20DC,than%20the%20totals%20from%20the

wetland_treatment_area_pct = (2+2*2)/100 # Percent required for the wetland according to 2016 Christenson 10 ways - very conservative. 0.5*2% is range with 2X being the maximum buffer size

# Choose the HUC to use for the important data
HUC_use = 'HUC6' #Choose the HUC level to group by: HUC6, HUC8, HUC10, HUC12

# Limit the output quotas to those allowed by the land
land_requirements = True
if land_requirements:
    land_HUC = HUC_use # Use the same HUC for all data
    land_calc = 'Percent' # Choose to calcluate what facilities can be treated: Percent (percent of each facility in the HUC) or LoadLow or LoadHigh (Only treat facilities below or above a certain load)
    # land_calc = 'LoadLow' # Choose to calcluate what facilities can be treated: Percent (percent of each facility in the HUC) or LoadLow or LoadHigh (Only treat facilities below or above a certain load)
    # land_calc = 'LoadHigh' # Choose to calcluate what facilities can be treated: Percent (percent of each facility in the HUC) or LoadLow or LoadHigh (Only treat facilities below or above a certain load)
else:
    land_HUC = 'None'
land_string = f'Land{land_HUC}-{land_calc}'

# Limit Gray nutrient treatment to match Green treatment
limit_gray = True # True or False
if limit_gray:
    limit_gray_string = 'GrayLimits'
else:
    limit_gray_string = 'NoGrayLimits'
    
# Uses combinations of green treatment methods
use_green_combinations = True
if use_green_combinations:
    green_combos_string = 'GreenCombos'
else:
    green_combos_string = 'NoGreenCombos'
    
# Use Gray Water Treatment Concentration Limitations
use_concentration_limits = True
if use_concentration_limits:
    concentration_limit_string = 'ConcLimits'
    n_conc_removal_limit = 32 # Assuming we use A20 Gray Treatment as our limiting factor 
    p_conc_removal_limit = 4.7 # Assuming we use A20 Gray Treatment as our limiting factor 
    max_treatment_cycles = 5 # Assume that we can treat this 5 times maximum with Gray
else:
    concentration_limit_string = 'NoConcLimits'

    
# Selects if we are inluding all N and P values, excluding the areas that have no treatment, or only including facilities with treatment of both nutrients
NP_output_type = 'NorPconc' # NandP, NorP, NPall, NorPconc

#%% Import Data
#Import data from the Hypoxia Task Force Nutrient Modeling - Includes DMR and Modeled loads
htf_filename = 'inputs/EPA Point Source Data/Nutrient-Modeling_Hypoxia_Task_Force_Search_clean.xlsx'
df_data_htf_raw = pd.read_excel(htf_filename,
                              skiprows=3,
                              dtype = {'HUC 12 Code': str,
                                       'FRS ID': str}
                              )

df_data_htf_raw.rename(columns={'HUC 12 Code': 'HUC12'}, inplace=True)

df_data_htf_raw.loc[df_data_htf_raw['HUC12'].str.len()<12,'HUC12']="0"+df_data_htf_raw['HUC12'] # Add leading zero if the HUC 12 code is only 11 characters

# Remove characters from HUC12 to get other huc values
df_data_htf_raw['HUC6'] = df_data_htf_raw['HUC12'].str[:-6]
df_data_htf_raw['HUC8'] = df_data_htf_raw['HUC12'].str[:-4]
df_data_htf_raw['HUC10'] = df_data_htf_raw['HUC12'].str[:-2]

# Set the value of 'Max Load' equal to Zero if 'Max Load' is NaN
df_data_htf_raw['Max Allowable Load Avalible'] = 'Y'
df_data_htf_raw['Max Allowable Load (lb/yr)'].fillna(0, inplace=True)
df_data_htf_raw.loc[df_data_htf_raw['Max Allowable Load (lb/yr)']==0,'Max Allowable Load Avalible'] = 'N'

# Set the value of 'Annual Flow' to the Average Facility Flow per day*365 if not NAN
df_data_htf_raw['Total Annual Flow (MGal/yr)'].fillna(df_data_htf_raw['Actual Average Facility Flow (MGD)']*365, inplace=True)

print('Raw Facility Data Imported!')

#%% Add EGRID to each Facility
print('Loading eGRID Data...')

# Using HUC12 becuase we want the finest resolution for each of the facilities. - Maybe it would be easier to just assign each facility an egrid based on their location. I should do that
huc_use_egrid = 'HUC12'

if huc_use_egrid == 'HUC8':
    egrid_filename = ''
elif huc_use_egrid == 'HUC12':
    egrid_filename = 'inputs/eGrid/HUC12_w_eGrid_v3_smalldf.pkl'

with open(egrid_filename, 'rb') as handle:
    egrid_HUC = pickle.load(handle)
    
HUCs = list(df_data_htf_raw[huc_use_egrid].unique())
HUCs.remove(HUCs[0])

##TODO: Make this a merge instead of a iteration - Just saves time, no change to the results
for HUC in HUCs:
    egrid_use_temp = egrid_HUC.loc[egrid_HUC[huc_use_egrid] == HUC].copy(deep=True)
    if huc_use_egrid == 'HUC8':
        egrid_use_temp.sort_values('overlap_area', axis=0, ascending=False, inplace=True)
        egrid_use_temp.reset_index(inplace=True)
        if len(egrid_use_temp) == 0:
            egrid_use = 'None'
        else:
            egrid_use = egrid_use_temp.loc[0,'ZipSubregi']
    elif huc_use_egrid == 'HUC12':
        egrid_use_temp.reset_index(inplace=True)
        if len(egrid_use_temp) == 0:
            egrid_use = 'None'
        else:
            egrid_use = egrid_use_temp.loc[0,'eGrid']
    df_data_htf_raw.loc[df_data_htf_raw[huc_use_egrid] == HUC, 'eGrid'] = egrid_use
      
print('eGRID Data Loaded!')
# asdfa
#%% Conversion Factors
kg_per_lb = 0.453592
L_per_gal = 3.78541

# Keep only data that has a more than a pound per year of nutrients and a total annual flow greater than 1 gallon
df_data_keep = df_data_htf_raw.loc[(df_data_htf_raw['Total Annual Flow (MGal/yr)']>1/1e6) & (df_data_htf_raw['Total Pounds (lb/yr)']>1)].copy(deep=True)

df_data_keep['Total Load (kg/yr)'] = df_data_keep['Total Pounds (lb/yr)'] * kg_per_lb
df_data_keep['Total Annual Flow (L/yr)'] = df_data_keep['Total Annual Flow (MGal/yr)'] * 1e6 * L_per_gal
df_data_keep['Mean Conc (mg/L)'] = df_data_keep['Total Load (kg/yr)'] * 1e6 / (df_data_keep['Total Annual Flow (L/yr)'])

df_data_N = df_data_keep.loc[(df_data_keep['Nutrient Type'] == 'Nitrogen')].copy(deep=True)
df_data_P = df_data_keep.loc[(df_data_keep['Nutrient Type'] == 'Phosphorus')].copy(deep=True)

# asD

#%% Import important HUC12 Data
print('Loading Important Data...')
with open('inputs/EPA EnviroAtlas/HUC12_Important_Data_v3_withAreasFixed.pkl', 'rb') as handle:
   df_importantData = pickle.load(handle)
   
# Remove characters from HUC12 to get other huc values
df_importantData['HUC6'] = df_importantData['HUC12'].str[:-6]
df_importantData['HUC8'] = df_importantData['HUC12'].str[:-4]
df_importantData['HUC10'] = df_importantData['HUC12'].str[:-2]

# Define a lambda function to compute the weighted mean:
wm = lambda x: np.average(x, weights=df_importantData.loc[x.index, "area_ha"])
# Lambda function to join strings and drop duplicates
# unique_str = lambda x: ', '.join(x.drop_duplicates())
unique_str = lambda x: x.drop_duplicates()

# Define how to aggregate various fields
agg_functions = { 
                 'area_ha': 'sum', 
                 'Cropland_pct': wm,       
                 'Pasture_pct': wm, 
                 'Wetlands_pct': wm, 
                 'StreamLength_km': 'sum', 
                 'StreamLength_TotalImpaired_km': 'sum', 
                 'StreamLength_NutImpaired_km': 'sum', 
                 'StreamLength_TempImpaired_km': 'sum', 
                 'Wastewater_AnnualFlow_mgy': 'sum', 
                 'TileDrain_Sum_m2': 'sum', 
                 'AgBuffer_pct': wm, 
                 'AgWetlandsPotential_pct': wm, 
                 'Fertilizer_kgN_ha_yr': wm,
                 'Fertilizer_kgP_ha_yr': wm,
                 'state': unique_str,
                 'Elec_$perkWh': wm,
                 'NoTill_pct': wm,
                 'CoverCrop_pct': wm,
                 'N_kg/ha': wm,
                 'P_kg/ha': wm,
                 'Cropland ($/ha)': wm,
                 'Pastureland ($/ha)': wm,
                 }

# Save original DF as raw DF
df_importantDataRaw = df_importantData.copy(deep=True)

# Create new DataFrame by combining rows with same id values
df_importantData = df_importantData.groupby([land_HUC]).aggregate(agg_functions)
df_importantData.reset_index(inplace=True)

print('Important Data Done!')

#%% Import Green and Grey Treatment Options

filename_inputs = 'inputs/model_inputs.xlsx'
# sheet_green = 'Green Water Treatment Illinois'
# sheet_green = 'Green Water Treatment IlliTest'
sheet_green = 'Green Water Treatment Illin0'
df_green = pd.read_excel(filename_inputs,
                                sheet_name = sheet_green,
                                index_col=1,
                                )

sheet_gray = 'Gray Water Treatment Options'
df_gray = pd.read_excel(filename_inputs,
                                sheet_name = sheet_gray,
                                index_col=1,
                                )

print('Importing All Data Done!')

#%% Find All Green Treatment Permentations
if use_green_combinations:
    green_land_options = df_green.loc[df_green['Type'] == 'Land'].index.tolist()
    green_barrier_options = df_green.loc[df_green['Type'] == 'Barrier'].index.tolist()
    
    # Find all the combinations of land treatement options
    green_land_options_comb = []
    combinations_max = len(green_land_options)
    for i in range(1,combinations_max+1):
        combs = itertools.combinations(green_land_options,i) 
        for option in list(combs): green_land_options_comb.append(list(option))
    
    # Add a single barrier option to each of the land treatment options combined
    green_all_options = green_land_options_comb.copy()
    for green_barrier_option in green_barrier_options:
        for green_land_option_comb in green_land_options_comb:
            green_all_options.append([green_barrier_option]+green_land_option_comb)
        
    # Remove all single options
    green_all_options = [x for x in green_all_options if len(x)>1]
            
    # Function to return column results cleaner
    def column_test(df,column_name):
        if len(df.loc[df[column_name] == 'Yes']) > 0:
            result = 'Yes'
        else:
            result = 'No'
        return result
    
    # Create new entry in the green df for each new combination
    for green_option in green_all_options:
        df_temp_option = df_green.loc[green_option].copy(deep=True)
        temp_name = '_'.join(df_temp_option['Abrev'].tolist())
        df_green.loc[temp_name,'Treatment Name'] = f'Combined: {", ".join(df_temp_option["Treatment Name"].tolist())}'
        df_green.loc[temp_name,'Abrev'] = temp_name
        df_green.loc[temp_name,'Type'] = 'Comb'
        df_green.loc[temp_name,'Level'] = df_green.loc[green_option,'Level'].min()
        df_green.loc[temp_name,'Green/Gray'] = 'Green'
        df_green.loc[temp_name,'2022Cost kgN'] = df_green.loc[green_option,'2022Cost kgN'].sum()
        df_green.loc[temp_name,'2022Cost kgP'] = df_green.loc[green_option,'2022Cost kgP'].sum()
        df_green.loc[temp_name,'GWP N'] = df_green.loc[green_option,'GWP N'].sum()
        df_green.loc[temp_name,'GWP P'] = df_green.loc[green_option,'GWP P'].sum()
        
        # Solve for new N Removal Percent
        n_removal_temp_list = df_green.loc[green_option,'N Removal'].tolist()
        n_removal_temp_list.sort(reverse = True)
        n_removal_temp = 1
        for pct in n_removal_temp_list: 
            if n_removal_temp == 1:
                n_removal_temp = pct
            else:
                n_removal_temp += pct * (1 - n_removal_temp)
            # print(pct,n_removal_temp)
        df_green.loc[temp_name,'N Removal'] = n_removal_temp
        
        # Solve for new P Removal Percent
        p_removal_temp_list = df_green.loc[green_option,'P Removal'].tolist()
        p_removal_temp_list.sort(reverse = True)
        p_removal_temp = 1
        for pct in p_removal_temp_list: 
            if p_removal_temp == 1:
                p_removal_temp = pct
            else:
                p_removal_temp += pct * (1 - p_removal_temp)
            # print(pct,p_removal_temp)
        df_green.loc[temp_name,'P Removal'] = p_removal_temp
            
        df_green.loc[temp_name,'eGrid Replace'] = column_test(df_temp_option,'eGrid Replace')
        df_green.loc[temp_name,'Tile Drain'] = column_test(df_temp_option,'Tile Drain')
        df_green.loc[temp_name,'Buffer'] = column_test(df_temp_option,'Buffer')
        df_green.loc[temp_name,'Wetlands'] = column_test(df_temp_option,'Wetlands')
        df_green.loc[temp_name,'Fertilizer'] = column_test(df_temp_option,'Fertilizer')
        df_green.loc[temp_name,'Farmer Incentive'] = column_test(df_temp_option,'Farmer Incentive')
        df_green.loc[temp_name,'Iowa Use'] = df_green.loc[green_option,'Iowa Use'].max()
        
print('Green Treatment Permentations Found!')

#%% Find Treatment Options
# Find all green treatment options
treatments_green = df_green.loc[df_green['Level']>=min_treat_level].index.tolist()
# treatments_green = df_green.index.tolist()
# treatments_green.remove(treatments_green[0]) # Remove nan value from units row

# Find all gray treatment options
treatments_gray = df_gray.loc[df_gray['Level']>=min_treat_level].index.tolist()
# treatments_gray.remove(treatments_gray[0]) # Remove nan value from units row
# treatments_gray = [treatments_gray[7],treatments_gray[8]]

#%% Adjust Emissions per EGRID Region
egrid_2010_us = pd.read_excel('inputs/eGrid/eGRID2010_Data.xls',
                             sheet_name = 'US10',
                             skiprows=4)
egrid_2010_us_col = 'USC2ERTA' # 'U.S. annual CO2 equivalent total output emission rate (lb/MWh)
egrid_2010_ghg_kgkwh = egrid_2010_us[egrid_2010_us_col].item() * 0.453592 / 1000 # Convert to kg/kwh


egrid_2021_all = pd.read_excel('inputs/eGrid/eGRID2021_data.xlsx',
                             sheet_name = 'SRL21',
                             skiprows=1)
egrid_2021_all.set_index('SUBRGN',inplace=True)
egrid_2021_all_col = 'SRC2ERTA' # eGRID subregion annual CO2 equivalent total output emission rate (lb/MWh)
egrid_2021_ghg_kgkwh = egrid_2021_all[egrid_2021_all_col] * 0.453592 / 1000 # Convert to kg/kwh

egrid_2021_us = pd.read_excel('inputs/eGrid/eGRID2021_data.xlsx',
                             sheet_name = 'US21',
                             skiprows=1)
egrid_2021_ghg_kgkwh['None'] = egrid_2021_us[egrid_2010_us_col].item() * 0.453592 / 1000 # Convert to kg/kwh

#%% Adjust costs per EIA updated data and geographic location

previous_elec_cost = 0.10 #$0.10/kWh was appropriate for use for this study based on the national average electricity price as of May 2014 (U.S. EIA, 2015). The 2014 electricity costs match the 2014dollar basis discussed in Section 3.2.1.
previous_elec_cost = previous_elec_cost *  1.24 # Convert from $2014 to $2022

#%% Treat some of the nutrients
def treat_nutrients(df_data_in_N, df_data_in_P, df_treatment, treat_conc_N, treat_conc_P):
    
    test_mode = False #TODO Change to Test Mode
    if test_mode:
        # treatment_green = 'Cover Crop'
        # treatment_green = 'Wetland'
        # treatment_green = 'No-till'
        treatment_green = 'Bioreactors'
        # treatment_green = 'Combined: Split N Application, Cover Crop'
        # treatment_green = 'N Rate Reduction'
        # treatment_green = 'W_NT'
        # treatment_green = 'BR_NR_NS_CC_NT'
        df_treatment = df_green.loc[treatment_green]
        print(f'Running in test moded with Green Treatment: {treatment_green}')
        
        treatment_gray = 'Level 2-1, A2O'
        # treatment_gray = 'Level 5-1, B5/RO'
        df_treatment = df_gray.loc[treatment_gray]
        print(f'Running in test moded with Gray Treatment: {treatment_gray}')
        
        df_data_in_N = df_data_N
        df_data_in_P = df_data_P
        
    
    df_data_in_N = df_data_in_N.copy(deep=True)
    df_data_in_P = df_data_in_P.copy(deep=True)
    
    df_data_in_N.set_index('NPDES Permit Number',inplace=True)
    df_data_in_P.set_index('NPDES Permit Number',inplace=True)
    
    n_permits = df_data_in_N.index.tolist()
    n_permits = [s for s in n_permits if not s.startswith('AK')] # Remove Alaska
    n_permits = [s for s in n_permits if not s.startswith('HI')] # Remove Hawaii
    n_permits = [s for s in n_permits if not s.startswith('PR')] # Remove Purto Rico
    n_permits = [s for s in n_permits if not s.startswith('AS')] # Remove American Samoa
    n_permits = [s for s in n_permits if not s.startswith('GU')] # Remove Guam
    n_permits = [s for s in n_permits if not s.startswith('VI')] # Remove Virgin Islands
    n_permits = [s for s in n_permits if not s.startswith('MP')] # Remove Mariana Islands
    
    p_permits = df_data_in_P.index.tolist()
    p_permits = [s for s in p_permits if not s.startswith('AK')] # Remove Alaska
    p_permits = [s for s in p_permits if not s.startswith('HI')] # Remove Hawaii
    p_permits = [s for s in p_permits if not s.startswith('PR')] # Remove Purto Rico
    p_permits = [s for s in p_permits if not s.startswith('AS')] # Remove American Samoa
    p_permits = [s for s in p_permits if not s.startswith('GU')] # Remove Guam
    p_permits = [s for s in p_permits if not s.startswith('VI')] # Remove Virgin Islands
    p_permits = [s for s in p_permits if not s.startswith('MP')] # Remove Mariana Islands
    
    all_permits = n_permits + p_permits
    all_permits = list(set(all_permits))
    # Filter out strings that start with 'PR'
    
    all_permits.sort()
    
    
    df_data_out = pd.DataFrame(index = all_permits) 
    
    df_data_out.loc[n_permits,'HUC6'] = df_data_in_N['HUC6']
    df_data_out.loc[p_permits,'HUC6'] = df_data_in_P['HUC6']
    
    df_data_out.loc[n_permits,'HUC8'] = df_data_in_N['HUC8']
    df_data_out.loc[p_permits,'HUC8'] = df_data_in_P['HUC8']
    
    df_data_out.loc[n_permits,'HUC10'] = df_data_in_N['HUC10']
    df_data_out.loc[p_permits,'HUC10'] = df_data_in_P['HUC10']
    
    df_data_out.loc[n_permits,'HUC12'] = df_data_in_N['HUC12']
    df_data_out.loc[p_permits,'HUC12'] = df_data_in_P['HUC12']
    
    df_data_out.loc[n_permits,'eGrid'] = df_data_in_N['eGrid']
    df_data_out.loc[p_permits,'eGrid'] = df_data_in_P['eGrid']
    
    # Filter out permits with no HUC or eGrid data
    df_data_out = df_data_out.dropna(how='all')
    
    # Solve for N Treatment Costs
    df_data_out['N Total Load (kg/yr)'] =  df_data_in_N['Total Load (kg/yr)']
    df_data_out['N Annual Flow (L/yr)'] =  df_data_in_N['Total Annual Flow (L/yr)']
    df_data_out['N Current Mean Conc (mg/L)'] =  df_data_in_N['Mean Conc (mg/L)']
    
    # Set the rest of the NaN values to zero
    df_data_out = df_data_out.fillna(0)

    if df_treatment['N Removal'] == 0:
        df_data_out['N Load Treated (kg/yr)'] = 0
    else:
        df_data_out['N Load Treated (kg/yr)'] = df_data_out['N Total Load (kg/yr)'] - df_data_out['N Annual Flow (L/yr)'] * treat_conc_N / 1e6
    
    # Remove Negative Values
    df_data_out.loc[df_data_out['N Load Treated (kg/yr)']<0,'N Load Treated (kg/yr)'] = 0
    
    # Adjust Nitrogen Treated based on geographic location
    if df_treatment['Tile Drain'] == 'Yes' and df_treatment['Green/Gray'] == 'Green':
        col_use_temp = 'TileDrain_Sum_m2'
        df_temp = df_data_out.reset_index().merge(df_importantData[[HUC_use,col_use_temp]], how="left", on=[HUC_use]).set_index('index')
        df_temp.loc[df_temp[col_use_temp]>0,col_use_temp] = 1
        df_temp[col_use_temp] = df_temp[col_use_temp].fillna(0)
        df_data_out['N Load Treated (kg/yr)'] = df_data_out['N Load Treated (kg/yr)'] * df_temp[col_use_temp]
     
    if df_treatment['Buffer'] == 'Yes' and df_treatment['Green/Gray'] == 'Green':
        col_use_temp = 'AgBuffer_pct'
        df_temp = df_data_out.reset_index().merge(df_importantData[[HUC_use,col_use_temp]], how="left", on=[HUC_use]).set_index('index')
        df_temp.loc[df_temp[col_use_temp]>0,col_use_temp] = 1
        df_temp[col_use_temp] = df_temp[col_use_temp].fillna(0)
        df_data_out['N Load Treated (kg/yr)'] = df_data_out['N Load Treated (kg/yr)'] * df_temp[col_use_temp]
      
    if df_treatment['Wetlands'] == 'Yes' and df_treatment['Green/Gray'] == 'Green':
        col_use_temp = 'AgWetlandsPotential_pct'
        col2_use_temp = 'Wetlands_pct'
        df_temp = df_data_out.reset_index().merge(df_importantData[[HUC_use,col_use_temp,col2_use_temp]], how="left", on=[HUC_use]).set_index('index')
        df_temp[[col_use_temp,col2_use_temp]] = df_temp[[col_use_temp,col2_use_temp]].fillna(0)
        df_temp.loc[(df_temp[col_use_temp]>0) | (df_temp[col2_use_temp]>0), 'wetland_use'] = 1
        df_temp.loc[(df_temp[col_use_temp]<=0) & (df_temp[col2_use_temp]<=0), 'wetland_use'] = 0
        df_data_out['N Load Treated (kg/yr)'] = df_data_out['N Load Treated (kg/yr)'] * df_temp['wetland_use']
        
    if df_treatment['Fertilizer'] == 'Yes' and df_treatment['Green/Gray'] == 'Green':
        col_use_temp = 'Fertilizer_kgN_ha_yr'
        df_temp = df_data_out.reset_index().merge(df_importantData[[HUC_use,col_use_temp]], how="left", on=[HUC_use]).set_index('index')
        df_temp.loc[df_temp[col_use_temp]>0,col_use_temp] = 1
        df_temp[col_use_temp] = df_temp[col_use_temp].fillna(0)
        df_data_out['N Load Treated (kg/yr)'] = df_data_out['N Load Treated (kg/yr)'] * df_temp[col_use_temp]
        
    if land_requirements and df_treatment['Green/Gray'] == 'Green':
        df_temp = df_data_out.reset_index()#.merge(df_importantData[[HUC_use,'area_ha','Cropland_pct','Pasture_pct', 'N_kg/ha', 'NoTill_pct', 'CoverCrop_pct']], how="left", on=[HUC_use]).set_index('index')
        
        # Define how to aggregate various fields
        agg_functions = { 
                         'N Total Load (kg/yr)': 'sum', 
                         'N Annual Flow (L/yr)': 'sum',       
                         'N Load Treated (kg/yr)': 'sum', 
                         }
        # Create new DataFrame by combining rows with same id values
        df_temp_HUC = df_temp.groupby([land_HUC]).aggregate(agg_functions)
        df_temp_HUC['N Current Mean Conc (mg/L)'] = df_temp_HUC['N Total Load (kg/yr)']*1e6/df_temp_HUC['N Annual Flow (L/yr)']
        df_temp_HUC = df_temp_HUC.reset_index().merge(df_importantData[[HUC_use,'area_ha','Cropland_pct','Pasture_pct', 'N_kg/ha', 'NoTill_pct', 'CoverCrop_pct']], how="left", on=[HUC_use]).set_index(HUC_use)
        
        #Need to make it so that the combined treatments also use cover crop and NT values for existing lands
        if ('Combined: ' in df_treatment['Treatment Name']) & (('No-till' in df_treatment['Treatment Name']) | ('Cover Crop' in df_treatment['Treatment Name'])):
            treatments_temp = df_treatment['Treatment Name'].replace('Combined: ', '').split(', ')
            
            df_temp_pct = pd.DataFrame(index = df_temp_HUC.index.tolist())
            for treatment_temp in treatments_temp:
                if treatment_temp == 'No-till':
                    df_temp_pct[treatment_temp] = df_temp_HUC['NoTill_pct']
                elif treatment_temp == 'Cover Crop':
                    df_temp_pct[treatment_temp] = df_temp_HUC['CoverCrop_pct']
                else: 
                    df_temp_pct[treatment_temp] = df_green.loc[treatment_temp,'Iowa Use']
                    
            # Find the maximum value in each row
            df_temp_pct['Use'] = df_temp_pct.max(axis=1)
            
        # Calculate maximum amount of ag land for treatment in each HUC 
        if df_treatment['Treatment Name'] == 'Cover Crop':
            df_temp_HUC['N Max HUC Treatment Area (ha)'] = df_temp_HUC['area_ha'] * (df_temp_HUC['Cropland_pct'] + df_temp_HUC['Pasture_pct']) * (1 - df_temp_HUC['CoverCrop_pct'])
    
        elif df_treatment['Treatment Name'] == 'No-till':
            df_temp_HUC['N Max HUC Treatment Area (ha)'] = df_temp_HUC['area_ha'] * (df_temp_HUC['Cropland_pct'] + df_temp_HUC['Pasture_pct']) * (1 - df_temp_HUC['NoTill_pct'])
        
        elif ('Combined: ' in df_treatment['Treatment Name']) & (('No-till' in df_treatment['Treatment Name']) | ('Cover Crop' in df_treatment['Treatment Name'])):
            df_temp_HUC['N Max HUC Treatment Area (ha)'] = df_temp_HUC['area_ha'] * (df_temp_HUC['Cropland_pct'] + df_temp_HUC['Pasture_pct']) * (1 - df_temp_pct['Use'])
        
        else:
            df_temp_HUC['N Max HUC Treatment Area (ha)'] = df_temp_HUC['area_ha'] * (df_temp_HUC['Cropland_pct'] + df_temp_HUC['Pasture_pct']) * (1 - df_treatment['Iowa Use'])
        
        # Calculate maximum amount of nutrient treatment possible in each HUC
        df_temp_HUC['N HUC Treatment Possible (kg/yr)'] = df_temp_HUC['N Max HUC Treatment Area (ha)'] * df_temp_HUC['N_kg/ha'] * df_treatment['N Removal'] 
        
        if land_calc == 'Percent':
            df_temp_HUC['Percent Load Wanted'] = df_temp_HUC['N Load Treated (kg/yr)']/df_temp_HUC['N HUC Treatment Possible (kg/yr)']
            df_temp_HUC['Percent Load Wanted'] = df_temp_HUC['Percent Load Wanted'].fillna(0) # Get rid of Nan values
            df_temp_HUC['Percent Use'] = 1
            df_temp_HUC.loc[df_temp_HUC['Percent Load Wanted'] > 1, 'Percent Use'] = 1/df_temp_HUC['Percent Load Wanted']
            df_temp_HUC.loc[df_temp_HUC['Percent Load Wanted'] == 0, 'Percent Use'] = 0
            df_temp_HUC.loc[df_temp_HUC['Percent Load Wanted'] < 1, 'Percent Use'] = 1
            
            df_temp_HUC.loc[df_temp_HUC['Percent Load Wanted'] > 1, 'N HUC Treatment Area (ha)'] = df_temp_HUC.loc[df_temp_HUC['Percent Load Wanted'] > 1, 'N Max HUC Treatment Area (ha)']
            df_temp_HUC.loc[df_temp_HUC['Percent Load Wanted'] <= 1, 'N HUC Treatment Area (ha)'] = df_temp_HUC.loc[df_temp_HUC['Percent Load Wanted'] <= 1, 'N Max HUC Treatment Area (ha)'] * df_temp_HUC.loc[df_temp_HUC['Percent Load Wanted'] <= 1, 'Percent Load Wanted']
            df_temp_HUC.loc[df_temp_HUC['Percent Load Wanted'] == 0, 'N HUC Treatment Area (ha)'] = 0
            df_temp_HUC['HUC Load Treated (kg/yr)'] = df_temp_HUC['N Load Treated (kg/yr)'] * df_temp_HUC['Percent Use']
            df_temp_HUC.reset_index(inplace = True)
        
            df_temp = df_temp.reset_index().merge(df_temp_HUC[[land_HUC,'Percent Use', 'N HUC Treatment Area (ha)', 'HUC Load Treated (kg/yr)']], how="left", on=[land_HUC]).set_index('index')
            df_temp['Percent Use'] = df_temp['Percent Use'].fillna(0)
            df_temp['N HUC Treatment Area (ha)'] = df_temp['N HUC Treatment Area (ha)'].fillna(0)
            df_temp['HUC Load Treated (kg/yr)'] = df_temp['HUC Load Treated (kg/yr)'].fillna(0)
            df_temp['New Load Treated (kg/yr)'] = df_temp['N Load Treated (kg/yr)'] * df_temp['Percent Use']
            df_temp['Area Needed (ha)'] = df_temp['New Load Treated (kg/yr)'] / df_temp['HUC Load Treated (kg/yr)'] * df_temp['N HUC Treatment Area (ha)']
            df_temp['Area Needed (ha)'] = df_temp['Area Needed (ha)'].fillna(0) # Get rid of Nan values
            
        ##TODO: Add area calcs for load options
        elif land_calc[:4] == 'Load':
            df_temp['New Load Treated (kg/yr)'] = 0
            for HUC_temp in df_temp_HUC.index.tolist():
                # HUC_temp = '03160204' # HUC for Testing HUC 8
                treatment_avalible = df_temp_HUC.loc[HUC_temp, 'N Treatment Possible (kg/yr)']
                if treatment_avalible == 0: continue
                df_temp_small = df_temp.loc[df_temp[land_HUC] == HUC_temp].copy(deep = True)
                if land_calc[4:] == 'High':  
                    df_temp_small.sort_values(by = 'N Load Treated (kg/yr)', ascending = False, inplace = True)
                elif land_calc[4:] == 'Low':
                    df_temp_small.sort_values(by = 'N Load Treated (kg/yr)', ascending = True, inplace = True)
                treatment_used = 0
                for name in df_temp_small.index.tolist():
                    if df_temp_small.loc[name,'N Load Treated (kg/yr)'] <= (treatment_avalible - treatment_used):
                        df_temp_small.loc[name,'New Load Treated (kg/yr)'] = df_temp_small.loc[name,'N Load Treated (kg/yr)']
                        df_temp.loc[name,'New Load Treated (kg/yr)'] = df_temp_small.loc[name,'N Load Treated (kg/yr)']
                        treatment_used += df_temp_small.loc[name,'N Load Treated (kg/yr)']
                    else:
                        df_temp.loc[name,'New Load Treated (kg/yr)'] = 0
                        
        df_data_out['N Load Treated (kg/yr)'] = df_temp['New Load Treated (kg/yr)']
        df_data_out['N Treatment Area (ha)'] = df_temp['Area Needed (ha)']
        
            
    # Limit Treatment Amounts for Gray Infrastructure
    if limit_gray and df_treatment['Green/Gray'] == 'Gray':
        df_data_out['N Load Treated (kg/yr)'] = df_gray_limits['N Load Treated (kg/yr)']
        
    df_data_out['N New Mean Conc (mg/L)'] =  (df_data_out['N Total Load (kg/yr)'] - df_data_out['N Load Treated (kg/yr)']) * 1e6 / df_data_out['N Annual Flow (L/yr)']
    df_data_out['N Conc Diff (mg/L)'] =  df_data_out['N Current Mean Conc (mg/L)'] - df_data_out['N New Mean Conc (mg/L)']
    
    #Fix Gray Concentration Limits
    if use_concentration_limits == True:
        df_data_out.loc[df_data_out['N Conc Diff (mg/L)'] > n_conc_removal_limit * max_treatment_cycles, 'N Load Treated (kg/yr)'] = 0
        df_data_out.loc[df_data_out['N Conc Diff (mg/L)'] > n_conc_removal_limit * max_treatment_cycles, 'N Treatment Area (ha)'] = 0
        df_data_out['N New Mean Conc (mg/L)'] =  (df_data_out['N Total Load (kg/yr)'] - df_data_out['N Load Treated (kg/yr)']) * 1e6 / df_data_out['N Annual Flow (L/yr)']
        df_data_out['N Conc Diff (mg/L)'] =  df_data_out['N Current Mean Conc (mg/L)'] - df_data_out['N New Mean Conc (mg/L)']
     
    
    # Calculate costs
    if (df_treatment['Green/Gray'] == 'Green') & (df_treatment['Farmer Incentive'] == 'Yes'):
        df_data_out['N Treated Cost ($/yr)'] = df_data_out['N Load Treated (kg/yr)'] * df_treatment['2022Cost kgN'] + df_data_out['N Treatment Area (ha)'] * farmer_incentive
        
        # Add land rental costs if wetland is used - need to do weighted average of land costs depending on the cropland to pastureland percentage split
        if 'Wetland' in df_treatment['Treatment Name']:
            df_temp = df_data_out.reset_index().merge(df_importantData[[HUC_use,'Cropland_pct','Pasture_pct','Cropland ($/ha)','Pastureland ($/ha)']], how="left", on=[HUC_use]).set_index('index')
            df_temp['area_pct_total'] = df_temp['Cropland_pct'] + df_temp['Pasture_pct']
            df_temp['area_cost_weighted'] = df_temp['Cropland_pct'] / df_temp['area_pct_total'] * df_temp['Cropland ($/ha)'] + df_temp['Pasture_pct'] / df_temp['area_pct_total'] * df_temp['Pastureland ($/ha)']
            df_temp['treated_land_cost'] = df_temp['N Treatment Area (ha)'] * df_temp['area_cost_weighted'] * wetland_treatment_area_pct
            df_temp['N Treated Cost ($/yr)'] = df_temp['N Treated Cost ($/yr)'] + df_temp['treated_land_cost']
            df_data_out['N Treated Cost ($/yr)'] = df_temp['N Treated Cost ($/yr)']
    
    elif (df_treatment['Green/Gray'] == 'Green') & (isinstance(df_treatment['Farmer Incentive'], float)):
        df_data_out['N Treated Cost ($/yr)'] = df_data_out['N Load Treated (kg/yr)'] * df_treatment['2022Cost kgN'] + df_data_out['N Treatment Area (ha)'] * farmer_incentive * df_treatment['Farmer Incentive']
        
        # Add land rental costs if wetland is used - need to do weighted average of land costs depending on the cropland to pastureland percentage split
        if 'Wetland' in df_treatment['Treatment Name']:
            df_temp = df_data_out.reset_index().merge(df_importantData[[HUC_use,'Cropland_pct','Pasture_pct','Cropland ($/ha)','Pastureland ($/ha)']], how="left", on=[HUC_use]).set_index('index')
            df_temp['area_pct_total'] = df_temp['Cropland_pct'] + df_temp['Pasture_pct']
            df_temp['area_cost_weighted'] = df_temp['Cropland_pct'] / df_temp['area_pct_total'] * df_temp['Cropland ($/ha)'] + df_temp['Pasture_pct'] / df_temp['area_pct_total'] * df_temp['Pastureland ($/ha)']
            df_temp['treated_land_cost'] = df_temp['N Treatment Area (ha)'] * df_temp['area_cost_weighted'] * wetland_treatment_area_pct
            df_temp['N Treated Cost ($/yr)'] = df_temp['N Treated Cost ($/yr)'] + df_temp['treated_land_cost']
            df_data_out['N Treated Cost ($/yr)'] = df_temp['N Treated Cost ($/yr)']
    
    elif (df_treatment['Green/Gray'] == 'Green') & (df_treatment['Farmer Incentive'] == 'No'):
        df_data_out['N Treated Cost ($/yr)'] = df_data_out['N Load Treated (kg/yr)'] * df_treatment['2022Cost kgN']
        # Add land rental costs if wetland is used
        if 'Wetland' in df_treatment['Treatment Name']:
            df_temp = df_data_out.reset_index().merge(df_importantData[[HUC_use,'Cropland_pct','Pasture_pct','Cropland ($/ha)','Pastureland ($/ha)']], how="left", on=[HUC_use]).set_index('index')
            df_temp['area_pct_total'] = df_temp['Cropland_pct'] + df_temp['Pasture_pct']
            df_temp['area_cost_weighted'] = df_temp['Cropland_pct'] / df_temp['area_pct_total'] * df_temp['Cropland ($/ha)'] + df_temp['Pasture_pct'] / df_temp['area_pct_total'] * df_temp['Pastureland ($/ha)']
            df_temp['treated_land_cost'] = df_temp['N Treatment Area (ha)'] * df_temp['area_cost_weighted'] * wetland_treatment_area_pct
            df_temp['N Treated Cost ($/yr)'] = df_temp['N Treated Cost ($/yr)'] + df_temp['treated_land_cost']
            df_data_out['N Treated Cost ($/yr)'] = df_temp['N Treated Cost ($/yr)']
        
    else:
        df_data_out['N Treated Cost ($/yr)'] = df_data_out['N Load Treated (kg/yr)'] * df_treatment['2022Cost kgN']
    
    df_data_out['N Treated GWP (tonnes-CO2eq/yr)'] = df_data_out['N Load Treated (kg/yr)'] * df_treatment['GWP N'] / 1000 # Convert to tonnes
    
    #Fix costs from Gray Concentration Limits
    if use_concentration_limits == True and df_treatment['Green/Gray'] == 'Gray':
        df_data_out.loc[df_data_out['N Conc Diff (mg/L)'] > n_conc_removal_limit, 'N Treated Cost ($/yr)'] = df_data_out.loc[df_data_out['N Conc Diff (mg/L)'] > n_conc_removal_limit, 'N Treated Cost ($/yr)'] * df_data_out.loc[df_data_out['N Conc Diff (mg/L)'] > n_conc_removal_limit, 'N Conc Diff (mg/L)'] / df_treatment['N Removal'] 
        df_data_out.loc[df_data_out['N Conc Diff (mg/L)'] > n_conc_removal_limit, 'N Treated GWP (tonnes-CO2eq/yr)'] = df_data_out.loc[df_data_out['N Conc Diff (mg/L)'] > n_conc_removal_limit, 'N Treated GWP (tonnes-CO2eq/yr)'] * df_data_out.loc[df_data_out['N Conc Diff (mg/L)'] > n_conc_removal_limit, 'N Conc Diff (mg/L)'] / df_treatment['N Removal']  
    
    # Solve for P Treatment Costs
    df_data_out['P Total Load (kg/yr)'] =  df_data_in_P['Total Load (kg/yr)']
    df_data_out['P Annual Flow (L/yr)'] =  df_data_in_P['Total Annual Flow (L/yr)']
    df_data_out['P Current Mean Conc (mg/L)'] =  df_data_in_P['Mean Conc (mg/L)']
    
    # Set NaN values for N Conc and P inputs to zero
    df_data_out = df_data_out.fillna(0)

    if df_treatment['P Removal'] == 0:
        df_data_out['P Load Treated (kg/yr)'] = 0
    else:
        df_data_out['P Load Treated (kg/yr)'] = df_data_out['P Total Load (kg/yr)'] - df_data_out['P Annual Flow (L/yr)'] * treat_conc_P / 1e6
    
    # Remove Negative Values
    df_data_out.loc[df_data_out['P Load Treated (kg/yr)']<0,'P Load Treated (kg/yr)'] = 0
    
    # Adjust Phosphous Treated based on geographic location
    if df_treatment['Tile Drain'] == 'Yes' and df_treatment['Green/Gray'] == 'Green':
        col_use_temp = 'TileDrain_Sum_m2'
        df_temp = df_data_out.reset_index().merge(df_importantData[[HUC_use,col_use_temp]], how="left", on=[HUC_use]).set_index('index')
        df_temp.loc[df_temp[col_use_temp]>0,col_use_temp] = 1
        df_temp[col_use_temp] = df_temp[col_use_temp].fillna(0)
        df_data_out['P Load Treated (kg/yr)'] = df_data_out['P Load Treated (kg/yr)'] * df_temp[col_use_temp]
     
    if df_treatment['Buffer'] == 'Yes' and df_treatment['Green/Gray'] == 'Green':
        col_use_temp = 'AgBuffer_pct'
        df_temp = df_data_out.reset_index().merge(df_importantData[[HUC_use,col_use_temp]], how="left", on=[HUC_use]).set_index('index')
        df_temp.loc[df_temp[col_use_temp]>0,col_use_temp] = 1
        df_temp[col_use_temp] = df_temp[col_use_temp].fillna(0)
        df_data_out['P Load Treated (kg/yr)'] = df_data_out['P Load Treated (kg/yr)'] * df_temp[col_use_temp]
      
    if df_treatment['Wetlands'] == 'Yes' and df_treatment['Green/Gray'] == 'Green':
        col_use_temp = 'AgWetlandsPotential_pct'
        col2_use_temp = 'Wetlands_pct'
        df_temp = df_data_out.reset_index().merge(df_importantData[[HUC_use,col_use_temp,col2_use_temp]], how="left", on=[HUC_use]).set_index('index')
        df_temp[[col_use_temp,col2_use_temp]] = df_temp[[col_use_temp,col2_use_temp]].fillna(0)
        df_temp.loc[(df_temp[col_use_temp]>0) | (df_temp[col2_use_temp]>0), 'wetland_use'] = 1
        df_temp.loc[(df_temp[col_use_temp]<=0) & (df_temp[col2_use_temp]<=0), 'wetland_use'] = 0
        df_data_out['P Load Treated (kg/yr)'] = df_data_out['P Load Treated (kg/yr)'] * df_temp['wetland_use']
        
    if df_treatment['Fertilizer'] == 'Yes' and df_treatment['Green/Gray'] == 'Green':
        col_use_temp = 'Fertilizer_kgN_ha_yr'
        df_temp = df_data_out.reset_index().merge(df_importantData[[HUC_use,col_use_temp]], how="left", on=[HUC_use]).set_index('index')
        df_temp.loc[df_temp[col_use_temp]>0,col_use_temp] = 1
        df_temp[col_use_temp] = df_temp[col_use_temp].fillna(0)
        df_data_out['P Load Treated (kg/yr)'] = df_data_out['P Load Treated (kg/yr)'] * df_temp[col_use_temp]
        
    if land_requirements and df_treatment['Green/Gray'] == 'Green':
        df_temp = df_data_out.reset_index()#.merge(df_importantData[[HUC_use,'area_ha','Cropland_pct','Pasture_pct', 'P_kg/ha', 'NoTill_pct', 'CoverCrop_pct']], how="left", on=[HUC_use]).set_index('index')
        # df_temp.loc[df_temp['area_ha']==0,'area_ha'] = 36*258.999 # Add mean value to each huc12 if area is zero. Source: https://www.epa.gov/rps/step-2-decide-which-watershed-units-screen
        
        # Define how to aggregate various fields
        agg_functions = { 
                         'P Total Load (kg/yr)': 'sum', 
                         'P Annual Flow (L/yr)': 'sum',       
                         'P Load Treated (kg/yr)': 'sum', 
                         }
        # Create new DataFrame by combining rows with same id values
        df_temp_HUC = df_temp.groupby([land_HUC]).aggregate(agg_functions)
        df_temp_HUC['P Current Mean Conc (mg/L)'] = df_temp_HUC['P Total Load (kg/yr)']*1e6/df_temp_HUC['P Annual Flow (L/yr)']
        df_temp_HUC = df_temp_HUC.reset_index().merge(df_importantData[[HUC_use,'area_ha','Cropland_pct','Pasture_pct', 'P_kg/ha', 'NoTill_pct', 'CoverCrop_pct']], how="left", on=[HUC_use]).set_index(HUC_use)
        
        
        # Calculate maximum amount of ag land for treatment in each HUC 
        if df_treatment['Treatment Name'] == 'Cover Crop':
            df_temp_HUC['P Max HUC Treatment Area (ha)'] = df_temp_HUC['area_ha'] * (df_temp_HUC['Cropland_pct'] + df_temp_HUC['Pasture_pct']) * (1 - df_temp_HUC['CoverCrop_pct'])
            
        elif df_treatment['Treatment Name'] == 'No-till':
            df_temp_HUC['P Max HUC Treatment Area (ha)'] = df_temp_HUC['area_ha'] * (df_temp_HUC['Cropland_pct'] + df_temp_HUC['Pasture_pct']) * (1 - df_temp_HUC['NoTill_pct'])
            
        elif ('Combined: ' in df_treatment['Treatment Name']) & (('No-till' in df_treatment['Treatment Name']) | ('Cover Crop' in df_treatment['Treatment Name'])):
            df_temp_HUC['P Max HUC Treatment Area (ha)'] = df_temp_HUC['area_ha'] * (df_temp_HUC['Cropland_pct'] + df_temp_HUC['Pasture_pct']) * (1 - df_temp_pct['Use'])
        
        else:
            df_temp_HUC['P Max HUC Treatment Area (ha)'] = df_temp_HUC['area_ha'] * (df_temp_HUC['Cropland_pct'] + df_temp_HUC['Pasture_pct']) * (1 - df_treatment['Iowa Use'])
            
        # Calculate maximum amount of nutrient treatment possible in each HUC
        df_temp_HUC['P HUC Treatment Possible (kg/yr)'] = df_temp_HUC['P Max HUC Treatment Area (ha)'] * df_temp_HUC['P_kg/ha'] * df_treatment['P Removal'] 
        
        if land_calc == 'Percent':
            df_temp_HUC['Percent Load Wanted'] = df_temp_HUC['P Load Treated (kg/yr)']/df_temp_HUC['P HUC Treatment Possible (kg/yr)']
            df_temp_HUC['Percent Load Wanted'] = df_temp_HUC['Percent Load Wanted'].fillna(0) # Get rid of Nan values
            df_temp_HUC['Percent Use'] = 1
            df_temp_HUC.loc[df_temp_HUC['Percent Load Wanted'] > 1, 'Percent Use'] = 1/df_temp_HUC['Percent Load Wanted']
            df_temp_HUC.loc[df_temp_HUC['Percent Load Wanted'] == 0, 'Percent Use'] = 0
            df_temp_HUC.loc[df_temp_HUC['Percent Load Wanted'] < 1, 'Percent Use'] = 1
            
            df_temp_HUC.loc[df_temp_HUC['Percent Load Wanted'] > 1, 'P HUC Treatment Area (ha)'] = df_temp_HUC.loc[df_temp_HUC['Percent Load Wanted'] > 1, 'P Max HUC Treatment Area (ha)']
            df_temp_HUC.loc[df_temp_HUC['Percent Load Wanted'] <= 1, 'P HUC Treatment Area (ha)'] = df_temp_HUC.loc[df_temp_HUC['Percent Load Wanted'] <= 1, 'P Max HUC Treatment Area (ha)'] * df_temp_HUC.loc[df_temp_HUC['Percent Load Wanted'] <= 1, 'Percent Load Wanted']
            df_temp_HUC.loc[df_temp_HUC['Percent Load Wanted'] == 0, 'P HUC Treatment Area (ha)'] = 0
            df_temp_HUC['HUC Load Treated (kg/yr)'] = df_temp_HUC['P Load Treated (kg/yr)'] * df_temp_HUC['Percent Use']
            df_temp_HUC.reset_index(inplace = True)
    
            df_temp = df_temp.reset_index().merge(df_temp_HUC[[land_HUC,'Percent Use', 'P HUC Treatment Area (ha)', 'HUC Load Treated (kg/yr)']], how="left", on=[land_HUC]).set_index('index')
            df_temp['Percent Use'] = df_temp['Percent Use'].fillna(0)
            df_temp['P HUC Treatment Area (ha)'] = df_temp['P HUC Treatment Area (ha)'].fillna(0)
            df_temp['HUC Load Treated (kg/yr)'] = df_temp['HUC Load Treated (kg/yr)'].fillna(0)
            df_temp['New Load Treated (kg/yr)'] = df_temp['P Load Treated (kg/yr)'] * df_temp['Percent Use']
            df_temp['Area Needed (ha)'] = df_temp['New Load Treated (kg/yr)'] / df_temp['HUC Load Treated (kg/yr)'] * df_temp['P HUC Treatment Area (ha)']
            df_temp['Area Needed (ha)'] = df_temp['Area Needed (ha)'].fillna(0) # Get rid of Nan values
        
        ##TODO: Add area calcs for load options    
        elif land_calc[:4] == 'Load':
            df_temp['New Load Treated (kg/yr)'] = 0
            for HUC_temp in df_temp_HUC.index.tolist():
                # HUC_temp = '03160204' # HUC for Testing HUC 8
                treatment_avalible = df_temp_HUC.loc[HUC_temp, 'P Treatment Possible (kg/yr)']
                if treatment_avalible == 0: continue
                df_temp_small = df_temp.loc[df_temp[land_HUC] == HUC_temp].copy(deep = True)
                if land_calc[4:] == 'High':  
                    df_temp_small.sort_values(by = 'P Load Treated (kg/yr)', ascending = False, inplace = True)
                elif land_calc[4:] == 'Low':
                    df_temp_small.sort_values(by = 'P Load Treated (kg/yr)', ascending = True, inplace = True)
                treatment_used = 0
                for name in df_temp_small.index.tolist():
                    if df_temp_small.loc[name,'P Load Treated (kg/yr)'] <= (treatment_avalible - treatment_used):
                        df_temp_small.loc[name,'New Load Treated (kg/yr)'] = df_temp_small.loc[name,'P Load Treated (kg/yr)']
                        df_temp.loc[name,'New Load Treated (kg/yr)'] = df_temp_small.loc[name,'P Load Treated (kg/yr)']
                        treatment_used += df_temp_small.loc[name,'P Load Treated (kg/yr)']
                    else:
                        df_temp.loc[name,'New Load Treated (kg/yr)'] = 0
        
        df_data_out['P Load Treated (kg/yr)'] = df_temp['New Load Treated (kg/yr)']
        df_data_out['P Treatment Area (ha)'] = df_temp['Area Needed (ha)']
    
    # Limit Treatment Amounts for Gray Infrastructure
    if limit_gray and df_treatment['Green/Gray'] == 'Gray':
        df_data_out['P Load Treated (kg/yr)'] = df_gray_limits['P Load Treated (kg/yr)']    
        
    df_data_out['P New Mean Conc (mg/L)'] =  (df_data_out['P Total Load (kg/yr)'] - df_data_out['P Load Treated (kg/yr)']) * 1e6 / df_data_out['P Annual Flow (L/yr)']
    df_data_out['P Conc Diff (mg/L)'] =  df_data_out['P Current Mean Conc (mg/L)'] - df_data_out['P New Mean Conc (mg/L)']
    
    #Input Gray Concentration Limits
    if use_concentration_limits == True:
        df_data_out.loc[df_data_out['P Conc Diff (mg/L)'] > p_conc_removal_limit * max_treatment_cycles, 'P Load Treated (kg/yr)'] = 0
        df_data_out.loc[df_data_out['P Conc Diff (mg/L)'] > p_conc_removal_limit * max_treatment_cycles, 'P Treatment Area (ha)'] = 0
        df_data_out['P New Mean Conc (mg/L)'] =  (df_data_out['P Total Load (kg/yr)'] - df_data_out['P Load Treated (kg/yr)']) * 1e6 / df_data_out['P Annual Flow (L/yr)']
        df_data_out['P Conc Diff (mg/L)'] =  df_data_out['P Current Mean Conc (mg/L)'] - df_data_out['P New Mean Conc (mg/L)']
    
    # Calculate Treatment Costs
    if (df_treatment['Green/Gray'] == 'Green') & (df_treatment['Farmer Incentive'] == 'Yes'):
        df_data_out['P Treated Cost ($/yr)'] = df_data_out['P Load Treated (kg/yr)'] * df_treatment['2022Cost kgP'] + df_data_out['P Treatment Area (ha)'] * farmer_incentive # I also tried converting to m3 for gray treatment but you get the same costs because I originally got the costs from the flow rates
    
        # Add land rental costs if wetland is used
        if 'Wetland' in df_treatment['Treatment Name']:
            df_temp = df_data_out.reset_index().merge(df_importantData[[HUC_use,'Cropland_pct','Pasture_pct','Cropland ($/ha)','Pastureland ($/ha)']], how="left", on=[HUC_use]).set_index('index')
            df_temp['area_pct_total'] = df_temp['Cropland_pct'] + df_temp['Pasture_pct']
            df_temp['area_cost_weighted'] = df_temp['Cropland_pct'] / df_temp['area_pct_total'] * df_temp['Cropland ($/ha)'] + df_temp['Pasture_pct'] / df_temp['area_pct_total'] * df_temp['Pastureland ($/ha)']
            df_temp['treated_land_cost'] = df_temp['P Treatment Area (ha)'] * df_temp['area_cost_weighted'] * wetland_treatment_area_pct
            df_temp['P Treated Cost ($/yr)'] = df_temp['P Treated Cost ($/yr)'] + df_temp['treated_land_cost']
            df_data_out['P Treated Cost ($/yr)'] = df_temp['P Treated Cost ($/yr)']
    
    elif (df_treatment['Green/Gray'] == 'Green') & (isinstance(df_treatment['Farmer Incentive'], float)):
        df_data_out['P Treated Cost ($/yr)'] = df_data_out['P Load Treated (kg/yr)'] * df_treatment['2022Cost kgP'] + df_data_out['P Treatment Area (ha)'] * farmer_incentive * df_treatment['Farmer Incentive']
        
        # Add land rental costs if wetland is used - need to do weighted average of land costs depending on the cropland to pastureland percentage split
        if 'Wetland' in df_treatment['Treatment Name']:
            df_temp = df_data_out.reset_index().merge(df_importantData[[HUC_use,'Cropland_pct','Pasture_pct','Cropland ($/ha)','Pastureland ($/ha)']], how="left", on=[HUC_use]).set_index('index')
            df_temp['area_pct_total'] = df_temp['Cropland_pct'] + df_temp['Pasture_pct']
            df_temp['area_cost_weighted'] = df_temp['Cropland_pct'] / df_temp['area_pct_total'] * df_temp['Cropland ($/ha)'] + df_temp['Pasture_pct'] / df_temp['area_pct_total'] * df_temp['Pastureland ($/ha)']
            df_temp['treated_land_cost'] = df_temp['P Treatment Area (ha)'] * df_temp['area_cost_weighted'] * wetland_treatment_area_pct
            df_temp['P Treated Cost ($/yr)'] = df_temp['P Treated Cost ($/yr)'] + df_temp['treated_land_cost']
            df_data_out['P Treated Cost ($/yr)'] = df_temp['P Treated Cost ($/yr)']
    
    elif (df_treatment['Green/Gray'] == 'Green') & (df_treatment['Farmer Incentive'] == 'No'):
        df_data_out['P Treated Cost ($/yr)'] = df_data_out['P Load Treated (kg/yr)'] * df_treatment['2022Cost kgP']
        
        # Add land rental costs if wetland is used
        if 'Wetland' in df_treatment['Treatment Name']:
            df_temp = df_data_out.reset_index().merge(df_importantData[[HUC_use,'Cropland_pct','Pasture_pct','Cropland ($/ha)','Pastureland ($/ha)']], how="left", on=[HUC_use]).set_index('index')
            df_temp['area_pct_total'] = df_temp['Cropland_pct'] + df_temp['Pasture_pct']
            df_temp['area_cost_weighted'] = df_temp['Cropland_pct'] / df_temp['area_pct_total'] * df_temp['Cropland ($/ha)'] + df_temp['Pasture_pct'] / df_temp['area_pct_total'] * df_temp['Pastureland ($/ha)']
            df_temp['treated_land_cost'] = df_temp['P Treatment Area (ha)'] * df_temp['area_cost_weighted'] * wetland_treatment_area_pct
            df_temp['P Treated Cost ($/yr)'] = df_temp['P Treated Cost ($/yr)'] + df_temp['treated_land_cost']
            df_data_out['P Treated Cost ($/yr)'] = df_temp['P Treated Cost ($/yr)']
    
    else:
        df_data_out['P Treated Cost ($/yr)'] = df_data_out['P Load Treated (kg/yr)'] * df_treatment['2022Cost kgP']
    
    df_data_out['P Treated GWP (tonnes-CO2eq/yr)'] = df_data_out['P Load Treated (kg/yr)'] * df_treatment['GWP P'] / 1000 # Convert to tonnes
    
    #Fix costs from Gray Concentration Limits
    if use_concentration_limits == True and df_treatment['Green/Gray'] == 'Gray':
        df_data_out.loc[df_data_out['P Conc Diff (mg/L)'] > p_conc_removal_limit, 'P Treated Cost ($/yr)'] = df_data_out.loc[df_data_out['P Conc Diff (mg/L)'] > p_conc_removal_limit, 'P Treated Cost ($/yr)'] * df_data_out.loc[df_data_out['P Conc Diff (mg/L)'] > p_conc_removal_limit, 'P Conc Diff (mg/L)'] / df_treatment['P Removal'] 
        df_data_out.loc[df_data_out['P Conc Diff (mg/L)'] > p_conc_removal_limit, 'P Treated GWP (tonnes-CO2eq/yr)'] = df_data_out.loc[df_data_out['P Conc Diff (mg/L)'] > p_conc_removal_limit, 'P Treated GWP (tonnes-CO2eq/yr)'] * df_data_out.loc[df_data_out['P Conc Diff (mg/L)'] > p_conc_removal_limit, 'P Conc Diff (mg/L)'] / df_treatment['P Removal']  
    
    # Set the NaN values for P Conc to zero
    df_data_out = df_data_out.fillna(0)
    
    # Fix eGrid Values and EIA Costs
    if df_treatment['eGrid Replace'] == 'Yes' and df_treatment['Green/Gray'] == 'Gray':
        for eGrid in list(df_data_out['eGrid'].unique()):
            gwp_new_m3 = df_treatment['GWP US Mix'] - df_treatment['Total Electricity Demand'] * egrid_2010_ghg_kgkwh + df_treatment['Total Electricity Demand'] * egrid_2021_ghg_kgkwh[eGrid]
            df_data_out.loc[df_data_out['eGrid'] == eGrid, 'N Treated GWP (tonnes-CO2eq/yr)'] = df_data_out['N Load Treated (kg/yr)'] * (gwp_new_m3 / (df_treatment['N Removal'] / 1000)) / 1000 # Convert to tonnes
            df_data_out.loc[df_data_out['eGrid'] == eGrid, 'P Treated GWP (tonnes-CO2eq/yr)'] = df_data_out['P Load Treated (kg/yr)'] * (gwp_new_m3 / (df_treatment['P Removal'] / 1000)) / 1000 # Convert to tonnes
          
        df_temp = df_data_out.reset_index().merge(df_importantDataRaw[['HUC12','area_ha','Elec_$perkWh']], how="left", on=['HUC12']).set_index('index')
        df_temp['New N Cost'] = df_treatment['2022Cost kgN'] - (df_treatment['Total Electricity Demand'] / df_treatment['N Removal'] * 1000) * previous_elec_cost + (df_treatment['Total Electricity Demand'] / df_treatment['N Removal'] * 1000) * df_temp['Elec_$perkWh']
        df_temp['New P Cost'] = df_treatment['2022Cost kgP'] - (df_treatment['Total Electricity Demand'] / df_treatment['P Removal'] * 1000) * previous_elec_cost + (df_treatment['Total Electricity Demand'] / df_treatment['P Removal'] * 1000) * df_temp['Elec_$perkWh']
        
        df_data_out['N Treated Cost ($/yr)'] = df_data_out['N Load Treated (kg/yr)'] * df_temp['New N Cost']
        df_data_out['P Treated Cost ($/yr)'] = df_data_out['P Load Treated (kg/yr)'] * df_temp['New P Cost']
        
    # Find the maximum cost and emissions from N or P treatment
    df_data_out['Max Treated Cost ($/yr)'] = df_data_out[['N Treated Cost ($/yr)', 'P Treated Cost ($/yr)']].max(axis=1)
    df_data_out['Max Treated GWP (tonnes-CO2eq/yr)'] = df_data_out[['N Treated GWP (tonnes-CO2eq/yr)', 'P Treated GWP (tonnes-CO2eq/yr)']].max(axis=1)
    
    # If the maximum cost and emissions is less than 0 but one of the methods isn't treated then choose the minimum
    df_data_out.loc[df_data_out['Max Treated Cost ($/yr)'] == 0, 'Max Treated Cost ($/yr)'] = df_data_out.loc[df_data_out['Max Treated Cost ($/yr)'] == 0, ['N Treated Cost ($/yr)', 'P Treated Cost ($/yr)']].min(axis=1)
    df_data_out.loc[df_data_out['Max Treated GWP (tonnes-CO2eq/yr)'] == 0, 'Max Treated GWP (tonnes-CO2eq/yr)'] = df_data_out.loc[df_data_out['Max Treated GWP (tonnes-CO2eq/yr)'] == 0,['N Treated GWP (tonnes-CO2eq/yr)', 'P Treated GWP (tonnes-CO2eq/yr)']].min(axis=1)
    
    
    # Save this as a global variable before filtering
    if df_treatment['Treatment Name'] == 'Cover Crop':
        global df_covercrop_facility
        df_covercrop_facility = df_data_out.copy(deep=True)
    
    if NP_output_type == 'NandP':
        df_data_out_filtered = df_data_out.loc[(df_data_out['N Load Treated (kg/yr)'] > 0) & (df_data_out['P Load Treated (kg/yr)'] > 0)].copy(deep=True)
    elif NP_output_type == 'NorP':
        df_data_out_filtered = df_data_out.loc[(df_data_out['N Load Treated (kg/yr)'] >= 0) | (df_data_out['P Load Treated (kg/yr)'] >= 0)].copy(deep=True)
    elif NP_output_type == 'NorPconc':
        df_data_out_filtered = df_data_out.loc[(df_data_out['N Current Mean Conc (mg/L)'] >= treat_conc_N) | (df_data_out['P Current Mean Conc (mg/L)'] >= treat_conc_P)].copy(deep=True)
    elif NP_output_type == 'NPall':
        df_data_out_filtered = df_data_out.copy(deep=True)
    
    # Find Summary Values
    df_summary = df_data_out_filtered.sum(axis=0)
    df_summary.drop(labels=['HUC6', 'HUC8', 'HUC10','HUC12','eGrid'], inplace=True)
    df_summary['Treatment'] = df_treatment['Treatment Name']
    df_summary['Treatment Type'] = df_treatment['Green/Gray']
    if df_data_out_filtered['N Annual Flow (L/yr)'].sum() == 0:
        df_summary['N Current Mean Conc (mg/L)'] = np.nan
        df_summary['N New Mean Conc (mg/L)'] = np.nan
        df_summary['N Conc Diff (mg/L)'] = np.nan
    else:
        df_summary['N Current Mean Conc (mg/L)'] = df_data_out_filtered['N Total Load (kg/yr)'].sum() * 1e6 / df_data_out_filtered['N Annual Flow (L/yr)'].sum()
        df_summary['N New Mean Conc (mg/L)'] = (df_data_out_filtered['N Total Load (kg/yr)'].sum() - df_data_out_filtered['N Load Treated (kg/yr)'].sum()) * 1e6 / df_data_out_filtered['N Annual Flow (L/yr)'].sum()
        df_summary['N Conc Diff (mg/L)'] = df_summary['N Current Mean Conc (mg/L)'] - df_summary['N New Mean Conc (mg/L)']
    
    if df_data_out_filtered['P Annual Flow (L/yr)'].sum() == 0:
        df_summary['P Current Mean Conc (mg/L)'] = np.nan
        df_summary['P New Mean Conc (mg/L)'] = np.nan
        df_summary['P Conc Diff (mg/L)'] = np.nan
    else:   
        df_summary['P Current Mean Conc (mg/L)'] = df_data_out_filtered['P Total Load (kg/yr)'].sum() * 1e6 / df_data_out_filtered['P Annual Flow (L/yr)'].sum()
        df_summary['P New Mean Conc (mg/L)'] = (df_data_out_filtered['P Total Load (kg/yr)'].sum() - df_data_out_filtered['P Load Treated (kg/yr)'].sum()) * 1e6 / df_data_out_filtered['P Annual Flow (L/yr)'].sum()
        df_summary['P Conc Diff (mg/L)'] = df_summary['P Current Mean Conc (mg/L)'] - df_summary['P New Mean Conc (mg/L)']

    return df_data_out_filtered, df_summary

# # Test the function
# treatment = 'Buffers'
# df_out_test,df_summary_test = treat_nutrients(df_data_N, df_data_P, df_green.loc[treatment], treat_conc_N, treat_conc_P)

#%% Run Anlaysis for Green Treatment Methods
print('Running Green Treatment Methods...')
dict_green_treatments = {}
for treatment_green in tqdm(treatments_green):
    dict_green_treatments[f'{treatment_green} facility'], dict_green_treatments[f'{treatment_green} Summary'] = treat_nutrients(df_data_N, df_data_P, df_green.loc[treatment_green], treat_conc_N, treat_conc_P)

print('Green Treatment Methods Done!')

#%% Find the optimum values for N and P treatment 
print('Running Green Treatment Optimization...')
green_options = df_green.iloc[1:,:].copy(deep=True)
green_options[['N Removal', 'P Removal']] = green_options[['N Removal', 'P Removal']].astype(float)
green_N_options = green_options.loc[green_options['N Removal'] > 0].index.tolist()
green_P_options = green_options.loc[green_options['P Removal'] > 0].index.tolist()

green_NP_combos = list(itertools.product(green_N_options, green_P_options))

# df_best_green_option = dict_green_treatments['Cover Crop facility'].copy(deep=True) # Pull this as the df template
df_best_green_option = df_covercrop_facility # Pull this as the df template
df_best_green_option.rename(columns = {'Max Treated Cost ($/yr)': 'Total Treated Cost ($/yr)',
                                       'Max Treated GWP (tonnes-CO2eq/yr)': 'Total Treated GWP (tonnes-CO2eq/yr)'
                                       }, inplace=True)
# asdfasdf
N_cols =[
     'N Load Treated (kg/yr)',
     'N Treatment Area (ha)',
     'N New Mean Conc (mg/L)',
     'N Conc Diff (mg/L)',
     'N Treated Cost ($/yr)',
     'N Treated GWP (tonnes-CO2eq/yr)',
     ]
# df_best_green_option[N_cols] = 0

P_cols =[
     'P Load Treated (kg/yr)',
     'P Treatment Area (ha)',
     'P New Mean Conc (mg/L)',
     'P Conc Diff (mg/L)',
     'P Treated Cost ($/yr)',
     'P Treated GWP (tonnes-CO2eq/yr)',
     ]
# df_best_green_option[P_cols] = 0
# df_best_green_option[['Total Treated Cost ($/yr)', 'Total Treated GWP (tonnes-CO2eq/yr)']] = 0
df_best_green_option[['N Option', 'P Option']] = 'Cover Crop'

green_NP_combos.remove(('Cover Crop', 'Cover Crop'))

dict_best_green_option = {
    'minimal_cost': df_best_green_option.copy(deep=True),
    'minimal_emissions': df_best_green_option.copy(deep=True),
    }

for green_NP_combo in tqdm(green_NP_combos):
    df_temp = dict_green_treatments[f'{green_NP_combo[0]} facility'][N_cols].copy(deep=True)
    df_temp[P_cols] = dict_green_treatments[f'{green_NP_combo[1]} facility'][P_cols]
    
    #If the technologies are the same then use the previously calculated values
    if green_NP_combo[0] == green_NP_combo[1]:
        temp_totalcost = dict_green_treatments[f'{green_NP_combo[0]} facility']['Max Treated Cost ($/yr)']
        temp_totalemissions = dict_green_treatments[f'{green_NP_combo[0]} facility']['Max Treated GWP (tonnes-CO2eq/yr)']
    
    #If the technologies are different then calculate new values
    else:
        option_n = green_NP_combo[0]
        if option_n in green_barrier_options + green_land_options:
            option_n = [df_green.loc[option_n, 'Abrev']]
        else:
            option_n = option_n.split('_')
        
        option_p = green_NP_combo[1]
        if option_p in green_barrier_options + green_land_options:
            option_p = [df_green.loc[option_p, 'Abrev']]
        else:
            option_p = option_p.split('_')
            
        option_both = list(set(option_n) & set(option_p))
        
        # If they have no common technologies then just add values together
        if len(option_both) == 0: 
            temp_totalcost = df_temp['N Treated Cost ($/yr)'] + df_temp['P Treated Cost ($/yr)']
            temp_totalemissions = df_temp['N Treated GWP (tonnes-CO2eq/yr)'] + df_temp['P Treated GWP (tonnes-CO2eq/yr)']
        
        # Else calculate the values for the individual parts
        else:
            # Find the technologies that are not shared
            option_n_only = [x for x in option_n if x not in option_both]
            option_p_only = [x for x in option_p if x not in option_both]
            
            temp_totalcosts2 = 0
            temp_totalemissions2 = 0
            use_farmer_incentive = False
            for option_temp in option_n_only:
                temp_totalcosts2 += df_green.loc[df_green['Abrev'] == option_temp, '2022Cost kgN'].item() * df_temp['N Load Treated (kg/yr)']
                temp_totalemissions2 += df_green.loc[df_green['Abrev'] == option_temp, 'GWP N'].item() * df_temp['N Load Treated (kg/yr)']
                
                if df_green.loc[df_green['Abrev'] == option_temp, 'Farmer Incentive'].item() == 'Yes': use_farmer_incentive = True
                
            for option_temp in option_p_only:
                temp_totalcosts2 += df_green.loc[df_green['Abrev'] == option_temp, '2022Cost kgP'].item() * df_temp['P Load Treated (kg/yr)']
                temp_totalemissions2 += df_green.loc[df_green['Abrev'] == option_temp, 'GWP P'].item() * df_temp['P Load Treated (kg/yr)']
                
                if df_green.loc[df_green['Abrev'] == option_temp, 'Farmer Incentive'].item() == 'Yes': use_farmer_incentive = True
                
            for option_temp in option_both:
                temp_cost_n = df_green.loc[df_green['Abrev'] == option_temp, '2022Cost kgN'].item() * df_temp['N Load Treated (kg/yr)']
                temp_cost_p = df_green.loc[df_green['Abrev'] == option_temp, '2022Cost kgP'].item() * df_temp['P Load Treated (kg/yr)']
                temp_emissions_n = df_green.loc[df_green['Abrev'] == option_temp, 'GWP N'].item() * df_temp['N Load Treated (kg/yr)']
                temp_emissions_p = df_green.loc[df_green['Abrev'] == option_temp, 'GWP P'].item() * df_temp['P Load Treated (kg/yr)']
                
                temp_totalcosts2 += pd.concat([temp_cost_n, temp_cost_p], axis=1).max(axis=1)
                temp_totalemissions2 += pd.concat([temp_emissions_n, temp_emissions_p], axis=1).max(axis=1)
                
                if df_green.loc[df_green['Abrev'] == option_temp, 'Farmer Incentive'].item() == 'Yes': use_farmer_incentive = True
             
            # Calculate the maximimum treatment area
            temp_maxarea2 = pd.concat([df_temp['N Treatment Area (ha)'], df_temp['P Treatment Area (ha)']], axis=1).max(axis=1)
            
            # Add land rental costs if wetland is used
            if 'W' in option_n_only + option_p_only + option_both:
                df_temp2 = dict_green_treatments['Cover Crop facility'][HUC_use].to_frame()
                unique_hucs = list(set(df_temp2[HUC_use].tolist()))
                
                ## TODO Make this matrix multipliation instead of a HUC loop
                for unique_huc in unique_hucs:
                    df_temp_ipdata = df_importantData.loc[df_importantData[HUC_use] == unique_huc, [HUC_use,'Cropland_pct','Pasture_pct','Cropland ($/ha)','Pastureland ($/ha)']]
                    if len(df_temp_ipdata) == 1:
                        # df_temp2.loc[df_temp2[HUC_use] == unique_huc, ['Cropland_pct','Pasture_pct','Cropland ($/ha)','Pastureland ($/ha)']] = df_temp_ipdata[['Cropland_pct','Pasture_pct','Cropland ($/ha)','Pastureland ($/ha)']]
                        df_temp2.loc[df_temp2[HUC_use] == unique_huc, 'Cropland_pct'] = df_temp_ipdata['Cropland_pct'].item()
                        df_temp2.loc[df_temp2[HUC_use] == unique_huc, 'Pasture_pct'] = df_temp_ipdata['Pasture_pct'].item()
                        df_temp2.loc[df_temp2[HUC_use] == unique_huc, 'Cropland ($/ha)'] = df_temp_ipdata['Cropland ($/ha)'].item()
                        df_temp2.loc[df_temp2[HUC_use] == unique_huc, 'Pastureland ($/ha)'] = df_temp_ipdata['Pastureland ($/ha)'].item()
                df_temp2 = df_temp2.fillna(0)

                df_temp2['area_pct_total'] = df_temp2['Cropland_pct'] + df_temp2['Pasture_pct']
                df_temp2['area_cost_weighted'] = df_temp2['Cropland_pct'] / df_temp2['area_pct_total'] * df_temp2['Cropland ($/ha)'] + df_temp2['Pasture_pct'] / df_temp2['area_pct_total'] * df_temp2['Pastureland ($/ha)']
                df_temp2 = df_temp2.fillna(0)
                
                temp_totalcosts2 += temp_maxarea2 * df_temp2['area_cost_weighted'] * wetland_treatment_area_pct
                
            # Set equal to the main cost and emissions values
            if use_farmer_incentive:
                temp_totalcost = temp_totalcosts2 + temp_maxarea2 * farmer_incentive
            else:
                temp_totalcost = temp_totalcosts2
                
            temp_totalemissions = temp_totalemissions2 / 1000 # Convert to tonnes
        
    for index_use in df_temp.index.tolist():
        
        if df_temp.loc[index_use,'N Load Treated (kg/yr)'] >= dict_best_green_option['minimal_cost'].loc[index_use,'N Load Treated (kg/yr)'] and df_temp.loc[index_use,'P Load Treated (kg/yr)'] >= dict_best_green_option['minimal_cost'].loc[index_use,'P Load Treated (kg/yr)']:
            if df_temp.loc[index_use,'N Load Treated (kg/yr)'] == 0 and df_temp.loc[index_use,'P Load Treated (kg/yr)'] == 0: 
               continue 
            
            # If one of the treated loads is higher, then we keep it. Always want the highest load
            if df_temp.loc[index_use,'N Load Treated (kg/yr)'] > dict_best_green_option['minimal_cost'].loc[index_use,'N Load Treated (kg/yr)'] or df_temp.loc[index_use,'P Load Treated (kg/yr)'] > dict_best_green_option['minimal_cost'].loc[index_use,'P Load Treated (kg/yr)']:
                dict_best_green_option['minimal_cost'].loc[index_use, N_cols] = df_temp.loc[index_use, N_cols]
                dict_best_green_option['minimal_cost'].loc[index_use, P_cols] = df_temp.loc[index_use, P_cols]
                dict_best_green_option['minimal_cost'].loc[index_use, 'Total Treated Cost ($/yr)'] = temp_totalcost.loc[index_use]
                dict_best_green_option['minimal_cost'].loc[index_use, 'Total Treated GWP (tonnes-CO2eq/yr)'] = temp_totalemissions.loc[index_use]
                dict_best_green_option['minimal_cost'].loc[index_use, 'N Option'] = green_NP_combo[0]
                dict_best_green_option['minimal_cost'].loc[index_use, 'P Option'] = green_NP_combo[1]
            
            # If the loads are equal, but the costs are less then we keep it.
            elif temp_totalcost.loc[index_use] < dict_best_green_option['minimal_cost'].loc[index_use, 'Total Treated Cost ($/yr)']:
                dict_best_green_option['minimal_cost'].loc[index_use, N_cols] = df_temp.loc[index_use, N_cols]
                dict_best_green_option['minimal_cost'].loc[index_use, P_cols] = df_temp.loc[index_use, P_cols]
                dict_best_green_option['minimal_cost'].loc[index_use, 'Total Treated Cost ($/yr)'] = temp_totalcost.loc[index_use]
                dict_best_green_option['minimal_cost'].loc[index_use, 'Total Treated GWP (tonnes-CO2eq/yr)'] = temp_totalemissions.loc[index_use]
                dict_best_green_option['minimal_cost'].loc[index_use, 'N Option'] = green_NP_combo[0]
                dict_best_green_option['minimal_cost'].loc[index_use, 'P Option'] = green_NP_combo[1]
         
        # Run for minimum emissions scenario    
        if df_temp.loc[index_use,'N Load Treated (kg/yr)'] >= dict_best_green_option['minimal_emissions'].loc[index_use,'N Load Treated (kg/yr)'] and df_temp.loc[index_use,'P Load Treated (kg/yr)'] >= dict_best_green_option['minimal_emissions'].loc[index_use,'P Load Treated (kg/yr)']:
            if df_temp.loc[index_use,'N Load Treated (kg/yr)'] == 0 and df_temp.loc[index_use,'P Load Treated (kg/yr)'] == 0: 
               continue
           
            # If one of the treated loads is higher, then we keep it. Always want the highest load
            if df_temp.loc[index_use,'N Load Treated (kg/yr)'] > dict_best_green_option['minimal_emissions'].loc[index_use,'N Load Treated (kg/yr)'] or df_temp.loc[index_use,'P Load Treated (kg/yr)'] > dict_best_green_option['minimal_emissions'].loc[index_use,'P Load Treated (kg/yr)']:
                dict_best_green_option['minimal_emissions'].loc[index_use, N_cols] = df_temp.loc[index_use, N_cols]
                dict_best_green_option['minimal_emissions'].loc[index_use, P_cols] = df_temp.loc[index_use, P_cols]
                dict_best_green_option['minimal_emissions'].loc[index_use, 'Total Treated Cost ($/yr)'] = temp_totalcost.loc[index_use]
                dict_best_green_option['minimal_emissions'].loc[index_use, 'Total Treated GWP (tonnes-CO2eq/yr)'] = temp_totalemissions.loc[index_use]
                dict_best_green_option['minimal_emissions'].loc[index_use, 'N Option'] = green_NP_combo[0]
                dict_best_green_option['minimal_emissions'].loc[index_use, 'P Option'] = green_NP_combo[1]
            
            # If the loads are equal, but the emissions are less then we keep it.
            elif temp_totalemissions.loc[index_use] < dict_best_green_option['minimal_emissions'].loc[index_use, 'Total Treated GWP (tonnes-CO2eq/yr)']:
                dict_best_green_option['minimal_emissions'].loc[index_use, N_cols] = df_temp.loc[index_use, N_cols]
                dict_best_green_option['minimal_emissions'].loc[index_use, P_cols] = df_temp.loc[index_use, P_cols]
                dict_best_green_option['minimal_emissions'].loc[index_use, 'Total Treated Cost ($/yr)'] = temp_totalcost.loc[index_use]
                dict_best_green_option['minimal_emissions'].loc[index_use, 'Total Treated GWP (tonnes-CO2eq/yr)'] = temp_totalemissions.loc[index_use]
                dict_best_green_option['minimal_emissions'].loc[index_use, 'N Option'] = green_NP_combo[0]
                dict_best_green_option['minimal_emissions'].loc[index_use, 'P Option'] = green_NP_combo[1]
   

if NP_output_type == 'NandP':
    dict_best_green_option['minimal_cost'] = dict_best_green_option['minimal_cost'].loc[(dict_best_green_option['minimal_cost']['N Load Treated (kg/yr)'] > 0) & (dict_best_green_option['minimal_cost']['P Load Treated (kg/yr)'] > 0)].copy(deep=True)
    dict_best_green_option['minimal_emissions'] = dict_best_green_option['minimal_emissions'].loc[(dict_best_green_option['minimal_emissions']['N Load Treated (kg/yr)'] > 0) & (dict_best_green_option['minimal_emissions']['P Load Treated (kg/yr)'] > 0)].copy(deep=True)
elif NP_output_type == 'NorP':
    dict_best_green_option['minimal_cost'] = dict_best_green_option['minimal_cost'].loc[(dict_best_green_option['minimal_cost']['N Load Treated (kg/yr)'] >= 0) | (dict_best_green_option['minimal_cost']['P Load Treated (kg/yr)'] >= 0)].copy(deep=True)
    dict_best_green_option['minimal_emissions'] = dict_best_green_option['minimal_emissions'].loc[(dict_best_green_option['minimal_emissions']['N Load Treated (kg/yr)'] >= 0) | (dict_best_green_option['minimal_emissions']['P Load Treated (kg/yr)'] >= 0)].copy(deep=True)
elif NP_output_type == 'NorPconc':
    dict_best_green_option['minimal_cost'] = dict_best_green_option['minimal_cost'].loc[(dict_best_green_option['minimal_cost']['N Current Mean Conc (mg/L)'] >= treat_conc_N) | (dict_best_green_option['minimal_cost']['P Current Mean Conc (mg/L)'] >= treat_conc_P)].copy(deep=True)
    dict_best_green_option['minimal_emissions'] = dict_best_green_option['minimal_emissions'].loc[(dict_best_green_option['minimal_emissions']['N Current Mean Conc (mg/L)'] >= treat_conc_N) | (dict_best_green_option['minimal_emissions']['P Current Mean Conc (mg/L)'] >= treat_conc_P)].copy(deep=True)
elif NP_output_type == 'NPall':
    dict_best_green_option['minimal_cost'] = dict_best_green_option['minimal_cost'].copy(deep=True)
    dict_best_green_option['minimal_emissions'] = dict_best_green_option['minimal_emissions'].copy(deep=True)

#Filter zero values so they have the same technology
names_list = ['minimal_cost', 'minimal_emissions']
for name in names_list:
    dict_best_green_option[name].loc[(dict_best_green_option[name]['N Load Treated (kg/yr)'] > 0) & (dict_best_green_option[name]['P Load Treated (kg/yr)'] == 0), 'P Option'] = dict_best_green_option[name]['N Option']
    dict_best_green_option[name].loc[(dict_best_green_option[name]['N Load Treated (kg/yr)'] == 0) & (dict_best_green_option[name]['P Load Treated (kg/yr)'] > 0), 'N Option'] = dict_best_green_option[name]['P Option']

#Add summaries for the best versions
def best_green_summaries(df_in, name):
    df_in.rename(columns = { 'Total Treated Cost ($/yr)': 'Max Treated Cost ($/yr)',
                                           'Total Treated GWP (tonnes-CO2eq/yr)': 'Max Treated GWP (tonnes-CO2eq/yr)'
                                           }, inplace=True)
    df_summary = df_in.sum(axis=0)
    df_summary.drop(labels=['HUC6', 'HUC8', 'HUC10','HUC12','eGrid', 'N Option', 'P Option'], inplace=True)
    df_summary['Treatment'] = name
    df_summary['Treatment Type'] = 'Green'
    if df_in['N Annual Flow (L/yr)'].sum() == 0:
        df_summary['N Current Mean Conc (mg/L)'] = np.nan
        df_summary['N New Mean Conc (mg/L)'] = np.nan
        df_summary['N Conc Diff (mg/L)'] = np.nan
    else:
        df_summary['N Current Mean Conc (mg/L)'] = df_in['N Total Load (kg/yr)'].sum() * 1e6 / df_in['N Annual Flow (L/yr)'].sum()
        df_summary['N New Mean Conc (mg/L)'] = (df_in['N Total Load (kg/yr)'].sum() - df_in['N Load Treated (kg/yr)'].sum()) * 1e6 / df_in['N Annual Flow (L/yr)'].sum()
        df_summary['N Conc Diff (mg/L)'] = df_summary['N Current Mean Conc (mg/L)'] - df_summary['N New Mean Conc (mg/L)']
        
    if df_in['P Annual Flow (L/yr)'].sum() == 0:
        df_summary['P Current Mean Conc (mg/L)'] = np.nan
        df_summary['P New Mean Conc (mg/L)'] = np.nan
        df_summary['P Conc Diff (mg/L)'] = np.nan
    else:   
        df_summary['P Current Mean Conc (mg/L)'] = df_in['P Total Load (kg/yr)'].sum() * 1e6 / df_in['P Annual Flow (L/yr)'].sum()
        df_summary['P New Mean Conc (mg/L)'] = (df_in['P Total Load (kg/yr)'].sum() - df_in['P Load Treated (kg/yr)'].sum()) * 1e6 / df_in['P Annual Flow (L/yr)'].sum()
        df_summary['P Conc Diff (mg/L)'] = df_summary['P Current Mean Conc (mg/L)'] - df_summary['P New Mean Conc (mg/L)']
               
    return df_summary

# Find the summary  of each optimial green dataset
for name in names_list:
    dict_best_green_option[f'{name}_summary'] = best_green_summaries(dict_best_green_option[name], name)

a_cost = dict_best_green_option['minimal_cost']
a_emissions = dict_best_green_option['minimal_emissions']

a_cost2 = a_cost.copy(deep=True)
a_emissions2 = a_emissions.copy(deep=True)

print('Green Treatment Optimization Done!')
# asdf

#%% Run Anlaysis for Gray Treatment Methods
print('Running Gray Treatment Methods...')
df_gray_limits = dict_best_green_option['minimal_cost']

dict_gray_treatments = {}
for treatment_gray in treatments_gray:
    dict_gray_treatments[f'{treatment_gray} facility'], dict_gray_treatments[f'{treatment_gray} Summary'] = treat_nutrients(df_data_N, df_data_P, df_gray.loc[treatment_gray], treat_conc_N, treat_conc_P)

print('Gray Treatment Methods Done!')
# asdf
#%% Create Summary DF for Comparison
print('Creating Summary DF for Comparison...')

# Run for Green Treatment Methods
df_summary_compare = pd.DataFrame()
for treatment_green in treatments_green:
    df_summary_compare = pd.concat([df_summary_compare, dict_green_treatments[f'{treatment_green} Summary'].to_frame().T], axis=0, ignore_index = True) 
    
# Run for Best Green Treatment Methods
for name in names_list:
    df_summary_compare = pd.concat([df_summary_compare, dict_best_green_option[f'{name}_summary'].to_frame().T], axis=0, ignore_index = True) 

# Run for Gray Treatment Methods
for treatment_gray in treatments_gray:
    df_summary_compare = pd.concat([df_summary_compare, dict_gray_treatments[f'{treatment_gray} Summary'].to_frame().T], axis=0, ignore_index = True) 

# Reset DF Index
df_summary_compare.reset_index(drop=True,inplace=True)

# Reorder the summary df columns
summary_cols = df_summary_compare.columns.tolist()
summary_cols = summary_cols[-2:] + summary_cols[:-2] # Move the last two columns to the first two columns
df_summary_compare = df_summary_compare[summary_cols] 

print('Summary DF Created!')

#%% Plot Histrograms
# ax = df_data_N['Mean Conc (mg/L)'].plot.hist(bins=12, alpha=0.5)
plot = False
if plot: 
    nutrient_type = 'Nitrogen'
    # nutrient_type = 'Phosphorus'
    if nutrient_type == 'Nitrogen':
        max_val = 50
        df = df_data_N.copy(deep=True)
    elif nutrient_type == 'Phosphorus':
        max_val = 5
        df = df_data_P.copy(deep=True)
    
    column = 'Mean Conc (mg/L)'
    df.loc[df[column]>max_val] = max_val
    
    plt.hist(df[column], bins=20)  # Adjust the number of bins as needed
    plt.xlabel(f'{nutrient_type} {column}')  # Set the x-axis label
    plt.ylabel('Frequency')  # Set the y-axis label
    plt.title(f'Histogram of {column}')  # Set the title of the plot
    plt.show()  # Display the plot

#%% Write the Files to Excel Files
print('Saving Excel Files...')

## TODO: Change Results Folder Name
todays_date = datetime.datetime.now().strftime('%Y-%m-%d')
folder_use = f'{todays_date} Level {min_treat_level}'
conc_str = f'point_source_{int(treat_conc_N)}concN_{int(treat_conc_P)}concP_postGeoChanges_{NP_output_type}_{land_string}_{green_combos_string}_{limit_gray_string}_{concentration_limit_string}'

conc_str = f'level-{min_treat_level}_{HUC_use}'


# Check if the results folder exists
if not os.path.exists(f'results/{folder_use}'):
    os.makedirs(f'results/{folder_use}') # Create the folder
    
with pd.ExcelWriter(f'results/{folder_use}/results_summaries_{conc_str}.xlsx') as writer_summaries:
    df_summary_compare.to_excel(writer_summaries, 
                                sheet_name='summary', 
                                index = False)
        
    

    
raw_results_filename = f'results/{folder_use}/results_facility_{conc_str}.xlsx'

with pd.ExcelWriter(raw_results_filename) as writer_raw:
    count = 0
    dict_keys = {}
    for treatment_green in treatments_green:
        count += 1
        dict_keys[str(count)] = f'{treatment_green}'
        dict_green_treatments[f'{treatment_green} facility'] = dict_green_treatments[f'{treatment_green} facility'].rename_axis('Facility')
        dict_green_treatments[f'{treatment_green} facility'].to_excel(writer_raw, sheet_name=str(count))
        
    for treatment_gray in treatments_gray:
        count += 1
        dict_keys[str(count)] = f'{treatment_gray}'
        dict_gray_treatments[f'{treatment_gray} facility'] = dict_gray_treatments[f'{treatment_gray} facility'].rename_axis('Facility')
        dict_gray_treatments[f'{treatment_gray} facility'].to_excel(writer_raw, sheet_name=str(count)) 
    
    count += 1
    dict_keys[str(count)] = 'Green Best Minimal Cost'
    dict_best_green_option['minimal_cost'] = dict_best_green_option['minimal_cost'].rename_axis('Facility')
    dict_best_green_option['minimal_cost'].to_excel(writer_raw, sheet_name=str(count))
    
    count += 1
    dict_keys[str(count)] = 'Green Best Minimal Emissions'
    dict_best_green_option['minimal_emissions'] = dict_best_green_option['minimal_emissions'].rename_axis('Facility')
    dict_best_green_option['minimal_emissions'].to_excel(writer_raw, sheet_name=str(count))
    
    df_keys = pd.DataFrame.from_dict(dict_keys, orient='index', columns=['Technology']).rename_axis('Sheet Name')
    df_keys.to_excel(writer_raw, sheet_name='keys')
    
## Move the keys sheet to the first sheet
# Load the Excel workbook
workbook = load_workbook(raw_results_filename)

# Get the index of the sheet named "keys"
index = workbook.sheetnames.index('keys')

# Move the sheet to the first position
workbook.move_sheet('keys', offset=-index)

# Save the modified workbook
workbook.save(raw_results_filename) 

print('Excel Files Saved!')

#%% Save Data as Pickle for Later Use
print('Saving Pickle...')
dict_results_all = {
    'dict_green_treatments': dict_green_treatments,
    'dict_gray_treatments': dict_gray_treatments,
    'df_summary_compare': df_summary_compare,
    'df_data_htf_raw': df_data_htf_raw,
    'df_data_keep': df_data_keep,
    'treat_conc_N': treat_conc_N,
    'treat_conc_P': treat_conc_P,
    'dict_best_green_option': dict_best_green_option,
    'df_importantData': df_importantData,
    'df_importantDataRaw': df_importantDataRaw,
    # 'mean_loss_N': mean_loss_N, 
    # 'mean_loss_P': mean_loss_P,
    'land_requirements': land_requirements,
    'land_HUC': land_HUC,
    'HUC_use': HUC_use,
    }

with open(f'results/{folder_use}/results_all_{conc_str}.pkl', 'wb') as handle:
    pickle.dump(dict_results_all, handle, protocol=pickle.HIGHEST_PROTOCOL) 
    
print('Pickle Saved!')

#%% End of Code        
execute_time = datetime.datetime.now() - begin_time
print('')
print('Code execution time: ', execute_time)